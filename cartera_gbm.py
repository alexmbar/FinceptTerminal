#!/usr/bin/env python3
"""
Lector de la cartera exportada desde GBM
=========================================

La app de GBM permite exportar el detalle del portafolio a Excel. El
archivo trae dos tablas en la misma hoja, una debajo de otra, sin mas
separacion que un renglon en blanco y el titulo de cada seccion en su
propio renglon:

    Mercado de Capitales Nacional
    Emisora/Fondo | Titulos | Costo promedio | ... | % Cartera
    DANHOS 13     | 15      | $29.39         | ... | 30.72%
    ...
    (renglon en blanco)
    Efectivo
    Emisora/Fondo | ...
    EFEC. MISMO DIA | ...

Este modulo solo lee el .xlsx y lo convierte a objetos Python. Cruzar cada
emisora contra el analisis de FIBRAs es trabajo de quien llama
(analizar_cbfi.py / reporte_cartera.py).

Requiere: pip install openpyxl
"""

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional


def _numero(valor) -> Optional[float]:
    """'$29.39', '-$9.72', '1.10%', '-' -> float o None."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if texto in ("", "-", "N/A"):
        return None
    negativo = texto.startswith("-")
    texto = texto.lstrip("-").replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        n = float(texto)
    except ValueError:
        return None
    return -n if negativo else n


def _porcentaje(valor) -> Optional[float]:
    """Igual que _numero, pero normaliza a fraccion: '30.72%' -> 0.3072."""
    n = _numero(valor)
    return n / 100 if n is not None else None


def normalizar_ticker(emisora: str) -> str:
    """'DANHOS 13' -> 'DANHOS13', para que calce con el CATALOGO de la BMV."""
    return re.sub(r"\s+", "", emisora or "").upper()


@dataclass
class PosicionCartera:
    """Un renglon de 'Mercado de Capitales Nacional': una emisora que tienes."""
    ticker: str
    nombre_gbm: str
    titulos: float
    costo_promedio: Optional[float] = None
    precio_mercado: Optional[float] = None
    valor_mercado: Optional[float] = None
    # Columna "P / M" de GBM: plusvalia o minusvalia en pesos de la
    # posicion completa (no por titulo). No incluye distribuciones cobradas.
    plusvalia_minusvalia: Optional[float] = None
    var_dia: Optional[float] = None
    pct_cartera: Optional[float] = None


@dataclass
class SaldoEfectivo:
    """Un renglon de 'Efectivo': una cuenta de liquidez (mismo dia, 24h, ...)."""
    cuenta: str
    valor_mercado: Optional[float] = None
    pct_cartera: Optional[float] = None


@dataclass
class Cartera:
    posiciones: list = field(default_factory=list)
    efectivo: list = field(default_factory=list)
    ruta: Optional[Path] = None

    @property
    def valor_posiciones(self) -> float:
        return sum(p.valor_mercado or 0 for p in self.posiciones)

    @property
    def valor_efectivo(self) -> float:
        return sum(e.valor_mercado or 0 for e in self.efectivo)

    @property
    def valor_total(self) -> float:
        return self.valor_posiciones + self.valor_efectivo

    @property
    def plusvalia_total(self) -> float:
        return sum(p.plusvalia_minusvalia or 0 for p in self.posiciones)


# El merge de celdas de GBM a veces recorta el texto del encabezado
# ("Emisora/F" en vez de "Emisora/Fondo"); se compara por prefijo.
_ENCABEZADOS = ("Emisora/Fondo", "Emisora/F")


def _es_fila_encabezado(primero: str) -> bool:
    return any(primero.startswith(e) for e in _ENCABEZADOS)


def leer_cartera(ruta) -> Cartera:
    """
    Lee el Excel que exporta la app de GBM (Detalle de Portafolio).

    Las dos tablas de la hoja se detectan por el titulo de su seccion
    ("Mercado de Capitales Nacional", "Efectivo"), no por un numero fijo de
    renglones: el tamano de la primera tabla cambia con cuantas emisoras
    tengas.
    """
    import openpyxl

    ruta = Path(ruta)
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[wb.sheetnames[0]]

    posiciones, efectivo = [], []
    seccion = None

    for fila in ws.iter_rows(values_only=True):
        primero = str(fila[0] or "").strip()
        if not primero:
            continue
        if primero.lower().startswith("mercado de capitales"):
            seccion = "posiciones"
            continue
        if primero.lower().startswith("efectivo"):
            seccion = "efectivo"
            continue
        if _es_fila_encabezado(primero) or seccion is None:
            continue
        # Titulo de una sub-seccion que no reconocemos por nombre (p. ej.
        # "Valores en Reporto"): no trae numero de titulos, a diferencia de
        # cualquier renglon de datos real (hasta el efectivo en $0 trae un
        # 0). Se ignora y la seccion vigente no cambia, así sus renglones
        # de datos se siguen sumando a la tabla en la que ya estábamos.
        if fila[1] is None:
            continue

        # Columnas: Emisora, Titulos, Costo prom, Precio mdo, PPP,
        # Valor mdo, P/M, %Var.Hist, %Var.Dia, ImpXCto, %Cartera
        if seccion == "posiciones":
            posiciones.append(PosicionCartera(
                ticker=normalizar_ticker(primero),
                nombre_gbm=primero,
                titulos=_numero(fila[1]) or 0,
                costo_promedio=_numero(fila[2]),
                precio_mercado=_numero(fila[3]),
                valor_mercado=_numero(fila[5]),
                plusvalia_minusvalia=_numero(fila[6]),
                var_dia=_porcentaje(fila[8]),
                pct_cartera=_porcentaje(fila[10]),
            ))
        else:
            efectivo.append(SaldoEfectivo(
                cuenta=primero,
                valor_mercado=_numero(fila[5]),
                pct_cartera=_porcentaje(fila[10]),
            ))

    return Cartera(posiciones=posiciones, efectivo=efectivo, ruta=ruta)
